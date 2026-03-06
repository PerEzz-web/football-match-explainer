import json
import html
import re
from datetime import datetime

import openpyxl
import streamlit as st
from openai import OpenAI

DEFAULT_ALLOWED_DOMAINS = [
    "fotmob.com",
    "sofascore.com",
    "flashscore.com",
    "whoscored.com",
    "soccerway.com",
    "fbref.com",
    "theanalyst.com",
    "transfermarkt.com",
    "espn.com",
    "onefootball.com",
]

BOOKMAKER_OPTIONS = {
    "Betlabel": {
        "name": "Betlabel",
        "url": "https://betlabel.com/en",
    },
    "Playzilla": {
        "name": "Playzilla",
        "url": "https://playzilla.com/en/",
    },
    "20Bet": {
        "name": "20Bet",
        "url": "https://20bet.com/",
    },
    "22Bet": {
        "name": "22Bet",
        "url": "https://22bet.com/",
    },
    "Ivibet": {
        "name": "Ivibet",
        "url": "https://ivibet.com/",
    },
    "N1 Bet": {
        "name": "N1 Bet",
        "url": "https://1n1bet.com/",
    },
}

PRESET_MODELS = [
    "gpt-5.4",
    "gpt-5",
    "gpt-5-mini",
    "gpt-5-nano",
    "gpt-4.1",
    "gpt-4.1-mini",
    "gpt-4o",
    "gpt-4o-mini",
    "Custom",
]

DEFAULT_SYSTEM_PROMPT = """You are an expert football analyst writing premium pre-match editorial copy for a betting/media website.

Your job is not to change the supplied forecast numbers. Your job is to explain and justify them in a way that feels like an expert match analysis written by a sharp football editor.

Core rules:
1. The forecast values from the Excel file are the source of truth.
2. Use recent, relevant football information from the web only to explain and support those values.
3. Stay aligned with the supplied forecast. Do not materially contradict it.
4. Never invent facts, lineups, injuries, suspensions, transfers, coaches, scorers, or news.
5. If some detail cannot be verified, do not guess. Briefly acknowledge uncertainty and continue with the strongest verified evidence.
6. Search only within the allowed domains provided by the app.
7. Use no more than the allowed number of tool calls.
8. Return JSON only and follow the exact schema.
9. Write plain text only. Do not output HTML tags.

What to look at first:
- the last five matches of each team
- the last five head-to-head matches between these teams, not only this season
- how strong each team has been in this specific competition or tournament
- goals scored and goals conceded
- home and away form if relevant
- recently signed players, important absences, and any new coach or major tactical change
- likely lineups for the upcoming match if available at the time of the request
- top scorers and key attacking players
- style of play: tempo, pressing, passes, shots, fouls, transitions, set-piece threat, defensive structure
- major team news that could affect the match
- anything else that could realistically influence the forecast

Writing style:
- Sound like a human football expert, not a machine.
- Write naturally, clearly, and confidently.
- Be user-friendly, but still analytical and specific.
- Add context, logic, and reasoning, not generic statements.
- Do not mention the model, the websites used, the search process, or links.
- Do not use citations, bullet lists, source notes, HTML, or markdown tables in the final text.
- Do not sound like a disclaimer.
- Do not promise certainty. Use smart, responsible language.

Very important evidence rules:
- Make the analysis feel grounded in real football evidence.
- In each section, when relevant, mention a few concrete anchors such as:
  - one recent result or form trend
  - one notable head-to-head pattern
  - one key player, scorer, creator, injury, suspension, or lineup point
  - one major tactical or coaching factor
  - one important recent news event affecting the match
- Do not overload the text with too many examples.
- Mention only the most relevant details that genuinely help justify the supplied forecast.
- It is better to mention 2-4 strong, specific anchors than many weak ones.
- Use player names, manager names, and notable match events when they matter.
- Do not list all matches or all events.

Very important consistency rules:
- All five output blocks must describe the same likely match story.
- The general match description must support the same scenario as the outcome, correct score, BTTS, and goals sections.
- Do not let one block imply an open high-scoring game while another implies a tight low-event match unless the forecast itself clearly supports that tension.
- The outcome explanation, correct score explanation, BTTS explanation, and goals explanation must be compatible with each other.
- If the most likely outcome is an away win, do not describe the home team as the likely dominant side unless clearly framed as a risk scenario.
- If BTTS leans No, do not describe both attacks as very likely to score freely.
- If BTTS leans Yes, do not describe one side as very unlikely to create chances unless you explain the contradiction carefully.
- If the most likely correct score is, for example, 0-2, the narrative should broadly fit that kind of match pattern.
- Before responding, do a final consistency check and rewrite any block that conflicts with the others.

Output goals:
- Explain how the match is likely to unfold.
- Explain why the supplied forecast numbers are reasonable.
- Mention what could make the prediction fail or become less reliable.
- Make the text ready to publish on a website with little or no editing.

Return these sections:
1. general_match_description
   - title
   - text
   - risk_note
2. match_outcome_probability
   - title
   - favored_outcome
   - text
3. correct_score_probability
   - title
   - most_likely_score
   - text
4. both_teams_to_score
   - title
   - most_likely_outcome
   - text
5. match_goals_probability
   - title
   - text

Content instructions:
- general_match_description: describe the likely game script, who may control the ball, who may create the better chances, whether the match should be open or controlled, and why. Include a few concrete details such as a recent result, a key player, or a relevant news item if they materially support the forecast. End with a concise note on what could break the prediction.
- match_outcome_probability: explain why the favored outcome has the edge using form, quality, match-up, competition context, and team news. Mention a few concrete supporting details when possible.
- correct_score_probability: explain the most likely exact score as the leading scenario among many possible outcomes, not as a certainty. Mention the most relevant concrete details that make that scoreline plausible.
- both_teams_to_score: explain whether both sides are likely to score based on attacking quality, defensive solidity, recent scoring patterns, and expected match state. Mention a few concrete anchors when useful.
- match_goals_probability: explain whether the profile points more toward a low-, medium-, or high-scoring game, using recent trends, tactical setup, and key team news. Mention a few concrete anchors when useful.

Final reminder:
The finished copy must read like expert football analysis for end users. It must be richer, more explicit, more insightful, and more concrete than a basic summary.
"""

OUTPUT_SCHEMA = {
    "type": "object",
    "properties": {
        "general_match_description": {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "text": {"type": "string"},
                "risk_note": {"type": "string"},
            },
            "required": ["title", "text", "risk_note"],
            "additionalProperties": False,
        },
        "match_outcome_probability": {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "favored_outcome": {"type": "string"},
                "text": {"type": "string"},
            },
            "required": ["title", "favored_outcome", "text"],
            "additionalProperties": False,
        },
        "correct_score_probability": {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "most_likely_score": {"type": "string"},
                "text": {"type": "string"},
            },
            "required": ["title", "most_likely_score", "text"],
            "additionalProperties": False,
        },
        "both_teams_to_score": {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "most_likely_outcome": {"type": "string"},
                "text": {"type": "string"},
            },
            "required": ["title", "most_likely_outcome", "text"],
            "additionalProperties": False,
        },
        "match_goals_probability": {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "text": {"type": "string"},
            },
            "required": ["title", "text"],
            "additionalProperties": False,
        },
    },
    "required": [
        "general_match_description",
        "match_outcome_probability",
        "correct_score_probability",
        "both_teams_to_score",
        "match_goals_probability",
    ],
    "additionalProperties": False,
}


def inject_css():
    st.markdown(
        """
        <style>
        .card {
            border: 1px solid rgba(49, 51, 63, 0.16);
            border-radius: 16px;
            padding: 16px 18px;
            background: #ffffff;
            color: #111827;
            margin-bottom: 12px;
        }
        .dark-card {
            border-radius: 18px;
            padding: 20px 22px;
            background: linear-gradient(135deg, #0f172a 0%, #18263d 100%);
            color: #ffffff;
            margin-bottom: 16px;
        }
        .section-title {
            font-size: 1.05rem;
            font-weight: 700;
            margin-bottom: 0.35rem;
            color: #111827;
        }
        .small-label {
            font-size: 0.82rem;
            color: #667085;
            margin-bottom: 0.35rem;
            text-transform: uppercase;
            letter-spacing: 0.04em;
        }
        .dark-card .small-label {
            color: rgba(255,255,255,0.75);
        }
        .big-number {
            font-size: 1.45rem;
            font-weight: 700;
            line-height: 1.2;
            color: #111827;
        }
        .muted-text {
            color: #475467;
            font-size: 0.94rem;
        }
        .pill {
            display: inline-block;
            padding: 5px 10px;
            border-radius: 999px;
            background: #eef2ff;
            color: #3730a3;
            font-size: 0.8rem;
            font-weight: 600;
            margin-top: 8px;
            margin-bottom: 8px;
        }
        .bookmaker-box {
            border: 1px solid rgba(49, 51, 63, 0.12);
            border-radius: 14px;
            padding: 14px 16px;
            background: #f8fafc;
            color: #111827;
            margin-bottom: 16px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def clean_text(value):
    if value is None:
        return None
    return str(value).replace("\xa0", " ").strip()


def clean_model_text(text):
    if text is None:
        return ""

    text = str(text)
    text = re.sub(r"<\s*br\s*/?\s*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</p\s*>", "\n\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<p\s*>", "", text, flags=re.IGNORECASE)
    text = re.sub(r"<li\s*>", "• ", text, flags=re.IGNORECASE)
    text = re.sub(r"</li\s*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = html.unescape(text)
    return text.strip()


def parse_percent(value):
    text = clean_text(value)
    if not text:
        return None
    text = text.replace("%", "").strip()
    try:
        return float(text) / 100
    except ValueError:
        return None


def format_pct(value):
    if value is None:
        return "—"
    return f"{value * 100:.1f}%"


def parse_match_datetime(value):
    text = clean_text(value)
    if not text:
        return None, None, None

    dt = None
    for fmt in ("%d-%m-%Y %H:%M", "%d/%m/%Y %H:%M"):
        try:
            dt = datetime.strptime(text, fmt)
            break
        except ValueError:
            pass

    if dt is None:
        return text, None, None

    return text, dt.strftime("%Y-%m-%d"), dt.strftime("%d %b %Y • %H:%M")


@st.cache_data
def load_matches_from_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    matches = []
    row = 2

    while row <= ws.max_row:
        top_left = clean_text(ws.cell(row, 1).value)
        bottom_left = clean_text(ws.cell(row + 1, 1).value) if row + 1 <= ws.max_row else None

        if top_left and ":" in top_left and bottom_left and " - " in bottom_left:
            raw_datetime, match_date_iso, match_date_display = parse_match_datetime(top_left)
            home_team, away_team = [part.strip() for part in bottom_left.split(" - ", 1)]

            home_win = parse_percent(ws.cell(row, 2).value)
            draw = parse_percent(ws.cell(row, 3).value)
            away_win = parse_percent(ws.cell(row, 4).value)

            correct_scores = [
                {
                    "score": clean_text(ws.cell(row, 5).value),
                    "probability": parse_percent(ws.cell(row + 1, 5).value),
                },
                {
                    "score": clean_text(ws.cell(row, 6).value),
                    "probability": parse_percent(ws.cell(row + 1, 6).value),
                },
                {
                    "score": clean_text(ws.cell(row, 7).value),
                    "probability": parse_percent(ws.cell(row + 1, 7).value),
                },
            ]

            btts_yes = parse_percent(ws.cell(row, 8).value)
            btts_no = parse_percent(ws.cell(row + 1, 8).value)

            totals = {
                "over_1_5": parse_percent(ws.cell(row, 9).value),
                "over_2_5": parse_percent(ws.cell(row, 10).value),
                "over_3_5": parse_percent(ws.cell(row, 11).value),
                "under_1_5": parse_percent(ws.cell(row, 12).value),
                "under_2_5": parse_percent(ws.cell(row, 13).value),
                "under_3_5": parse_percent(ws.cell(row, 14).value),
            }

            handicaps = {
                "Home +0.5": parse_percent(ws.cell(row, 15).value),
                "Home +1.5": parse_percent(ws.cell(row, 16).value),
                "Home +2.5": parse_percent(ws.cell(row, 17).value),
                "Away +0.5": parse_percent(ws.cell(row, 18).value),
                "Away +1.5": parse_percent(ws.cell(row, 19).value),
                "Away +2.5": parse_percent(ws.cell(row, 20).value),
                "Home -0.5": parse_percent(ws.cell(row, 21).value),
                "Home -1.5": parse_percent(ws.cell(row, 22).value),
                "Home -2.5": parse_percent(ws.cell(row, 23).value),
                "Away -0.5": parse_percent(ws.cell(row, 24).value),
                "Away -1.5": parse_percent(ws.cell(row, 25).value),
                "Away -2.5": parse_percent(ws.cell(row, 26).value),
            }

            favored_outcome = max(
                [("Home win", home_win), ("Draw", draw), ("Away win", away_win)],
                key=lambda item: item[1] if item[1] is not None else -1,
            )[0]

            most_likely_score = max(
                correct_scores,
                key=lambda item: item["probability"] if item["probability"] is not None else -1,
            )["score"]

            most_likely_btts = "Yes" if (btts_yes or 0) >= (btts_no or 0) else "No"

            matches.append(
                {
                    "label": f"{home_team} vs {away_team} — {match_date_display or raw_datetime}",
                    "home_team": home_team,
                    "away_team": away_team,
                    "match_date": match_date_iso or raw_datetime,
                    "match_date_display": match_date_display or raw_datetime,
                    "engine_forecast": {
                        "match_outcome_probability": {
                            "home_win": home_win,
                            "draw": draw,
                            "away_win": away_win,
                            "favored_outcome": favored_outcome,
                        },
                        "correct_score_probability": {
                            "top_outcomes": correct_scores,
                            "most_likely_score": most_likely_score,
                        },
                        "both_teams_to_score": {
                            "yes": btts_yes,
                            "no": btts_no,
                            "most_likely_outcome": most_likely_btts,
                        },
                        "match_goals_probability": totals,
                        "handicaps": handicaps,
                    },
                }
            )

            row += 2
        else:
            row += 1

    return matches


def build_user_payload(match, bookmaker):
    return {
        "match": {
            "home_team": match["home_team"],
            "away_team": match["away_team"],
            "match_date": match["match_date"],
            "output_language": "en",
        },
        "partner_bookmaker": {
            "name": bookmaker["name"],
            "url": bookmaker["url"],
            "placeholder_odds": "XX",
        },
        "engine_forecast": match["engine_forecast"],
    }


def format_goals_market_key(key):
    parts = key.split("_")
    if len(parts) == 3:
        side = parts[0].capitalize()
        line = f"{parts[1]}.{parts[2]}"
        return f"{side} {line}"
    return key


def get_top_goals_market(goals_data):
    valid_items = {k: v for k, v in goals_data.items() if v is not None}
    if not valid_items:
        return "Goals market"
    top_key = max(valid_items, key=valid_items.get)
    return format_goals_market_key(top_key)


def append_bookmaker_note(base_text, note):
    base_text = clean_model_text(base_text).strip()
    note = note.strip()

    if not base_text:
        return note

    if base_text[-1] not in ".!?":
        base_text += "."

    return f"{base_text} {note}"


def make_bookmaker_note(section_key, bookmaker, match):
    name = bookmaker["name"]
    url = bookmaker["url"]
    odds_link = f"[XX]({url})"

    forecast = match["engine_forecast"]

    favored_outcome = forecast["match_outcome_probability"]["favored_outcome"]
    likely_score = forecast["correct_score_probability"]["most_likely_score"]
    btts_outcome = forecast["both_teams_to_score"]["most_likely_outcome"]
    top_goals_market = get_top_goals_market(forecast["match_goals_probability"])

    if section_key == "general_match_description":
        return (
            f"{name} broadly leans towards that overall match script in this demo "
            f"and offers {odds_link} for that angle."
        )

    if section_key == "match_outcome_probability":
        return (
            f"{name} also leans towards a {favored_outcome.lower()} in this demo "
            f"and lists {odds_link} for that outcome."
        )

    if section_key == "correct_score_probability":
        return (
            f"{name} also prices the {likely_score} scoreline at {odds_link} in this demo."
        )

    if section_key == "both_teams_to_score":
        return (
            f"{name} also leans towards BTTS {btts_outcome.lower()} in this demo "
            f"and shows {odds_link} for that market."
        )

    if section_key == "match_goals_probability":
        return (
            f"{name} also points towards {top_goals_market} in this demo "
            f"and offers {odds_link} for that line."
        )

    return f"{name} also shows {odds_link} for this market in the demo."


def generate_explanation(match, bookmaker, model_name, system_prompt, allowed_domains, max_tool_calls):
    client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])

    response = client.responses.create(
        model=model_name,
        instructions=system_prompt,
        input=json.dumps(build_user_payload(match, bookmaker), ensure_ascii=False),
        tools=[
            {
                "type": "web_search",
                "filters": {"allowed_domains": allowed_domains},
            }
        ],
        include=["web_search_call.action.sources"],
        max_tool_calls=max_tool_calls,
        text={
            "format": {
                "type": "json_schema",
                "name": "kickform_match_explanation",
                "strict": True,
                "schema": OUTPUT_SCHEMA,
            }
        },
    )

    return json.loads(response.output_text)


def render_info_card(label, value, subtitle=None):
    safe_label = html.escape(str(label))
    safe_value = html.escape(str(value))
    safe_subtitle = html.escape(str(subtitle)) if subtitle else ""

    st.markdown(
        f"""
        <div class="card">
            <div class="small-label">{safe_label}</div>
            <div class="big-number">{safe_value}</div>
            {"<div class='muted-text'>" + safe_subtitle + "</div>" if subtitle else ""}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_match_header(match):
    home_team = html.escape(match["home_team"])
    away_team = html.escape(match["away_team"])
    match_date_display = html.escape(str(match["match_date_display"]))

    st.markdown(
        f"""
        <div class="dark-card">
            <div class="small-label">Selected match</div>
            <div style="font-size: 1.9rem; font-weight: 800; line-height: 1.2;">
                {home_team} vs {away_team}
            </div>
            <div style="margin-top: 8px; font-size: 1rem; color: rgba(255,255,255,0.85);">
                {match_date_display}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_bookmaker_box(bookmaker):
    st.markdown(
        f"""
        <div class="bookmaker-box">
            <div class="small-label">Partner bookmaker</div>
            <div style="font-size: 1.05rem; font-weight: 700;">
                <a href="{html.escape(bookmaker["url"])}" target="_blank">{html.escape(bookmaker["name"])}</a>
            </div>
            <div class="muted-text" style="margin-top: 6px;">
                Demo mode: analysis blocks will include a clickable placeholder odds link using XX.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_match_data(match):
    forecast = match["engine_forecast"]

    st.subheader("Match data")
    render_match_header(match)

    st.markdown("#### Match outcome probabilities")
    col1, col2, col3 = st.columns(3)
    with col1:
        render_info_card("Home win", format_pct(forecast["match_outcome_probability"]["home_win"]))
    with col2:
        render_info_card("Draw", format_pct(forecast["match_outcome_probability"]["draw"]))
    with col3:
        render_info_card("Away win", format_pct(forecast["match_outcome_probability"]["away_win"]))

    st.markdown("#### Most likely correct scores")
    score_cols = st.columns(3)
    for idx, item in enumerate(forecast["correct_score_probability"]["top_outcomes"]):
        with score_cols[idx]:
            render_info_card(item["score"] or "—", format_pct(item["probability"]), "Probability")

    st.markdown("#### Both teams to score")
    col1, col2 = st.columns(2)
    with col1:
        render_info_card("Yes", format_pct(forecast["both_teams_to_score"]["yes"]))
    with col2:
        render_info_card("No", format_pct(forecast["both_teams_to_score"]["no"]))

    st.markdown("#### Goals markets")
    row1 = st.columns(3)
    row2 = st.columns(3)

    with row1[0]:
        render_info_card("Over 1.5", format_pct(forecast["match_goals_probability"]["over_1_5"]))
    with row1[1]:
        render_info_card("Over 2.5", format_pct(forecast["match_goals_probability"]["over_2_5"]))
    with row1[2]:
        render_info_card("Over 3.5", format_pct(forecast["match_goals_probability"]["over_3_5"]))

    with row2[0]:
        render_info_card("Under 1.5", format_pct(forecast["match_goals_probability"]["under_1_5"]))
    with row2[1]:
        render_info_card("Under 2.5", format_pct(forecast["match_goals_probability"]["under_2_5"]))
    with row2[2]:
        render_info_card("Under 3.5", format_pct(forecast["match_goals_probability"]["under_3_5"]))

    with st.expander("Show handicap probabilities"):
        handicap_items = list(forecast["handicaps"].items())
        for start in range(0, len(handicap_items), 3):
            cols = st.columns(3)
            for col, (label, value) in zip(cols, handicap_items[start:start + 3]):
                with col:
                    render_info_card(label, format_pct(value))


def render_analysis_block(title, text, badge_text=None, risk_note=None):
    with st.container(border=True):
        st.markdown(f"#### {clean_model_text(title)}")

        if badge_text:
            safe_badge = html.escape(clean_model_text(badge_text))
            st.markdown(
                f'<div class="pill">{safe_badge}</div>',
                unsafe_allow_html=True,
            )

        st.markdown(clean_model_text(text))

        if risk_note:
            st.warning(f"**What could go wrong:** {clean_model_text(risk_note)}")


st.set_page_config(page_title="Kickform LLM explainer", layout="wide")
inject_css()

if "last_result" not in st.session_state:
    st.session_state["last_result"] = None
if "last_result_key" not in st.session_state:
    st.session_state["last_result_key"] = None

st.title("Kickform LLM explainer")
st.caption(
    "Select a match from your forecast file, review the forecast data, choose a partner bookmaker, then generate a readable expert-style explanation."
)

with st.sidebar:
    st.header("Settings")

    model_choice = st.selectbox("GPT model", PRESET_MODELS, index=0)
    if model_choice == "Custom":
        model_name = st.text_input("Custom model ID", value="gpt-5.4")
    else:
        model_name = model_choice

    partner_bookmaker_key = st.selectbox(
        "Partner bookmaker",
        list(BOOKMAKER_OPTIONS.keys()),
        index=0,
    )
    selected_bookmaker = BOOKMAKER_OPTIONS[partner_bookmaker_key]

    max_tool_calls = st.slider("Max web search calls", 1, 10, 4)

    domains_text = st.text_area(
        "Allowed websites (one per line)",
        value="\n".join(DEFAULT_ALLOWED_DOMAINS),
        height=180,
    )
    allowed_domains = [line.strip() for line in domains_text.splitlines() if line.strip()]

    system_prompt = st.text_area(
        "System prompt",
        value=DEFAULT_SYSTEM_PROMPT,
        height=500,
    )

matches = load_matches_from_excel("Forecasts.xlsx")

if not matches:
    st.error("No matches were found in Forecasts.xlsx")
    st.stop()

selected_label = st.selectbox("Select a match", [m["label"] for m in matches])
selected_match = next(m for m in matches if m["label"] == selected_label)

current_result_key = f"{selected_label}__{partner_bookmaker_key}__{model_name}"

render_match_data(selected_match)
render_bookmaker_box(selected_bookmaker)

if "OPENAI_API_KEY" not in st.secrets:
    st.error("OPENAI_API_KEY is missing from your Streamlit secrets.")
    st.stop()

if st.button("Generate explanation", type="primary"):
    try:
        with st.status("Working on your explanation", expanded=True) as status:
            st.write("Step 1/5 — Reading the selected match and forecast values from the Excel file.")
            payload = build_user_payload(selected_match, selected_bookmaker)

            st.write("Step 2/5 — Preparing the research brief, bookmaker context, and app settings.")
            _ = {
                "model": model_name,
                "allowed_domains_count": len(allowed_domains),
                "max_tool_calls": max_tool_calls,
                "payload_ready": bool(payload),
                "bookmaker": selected_bookmaker["name"],
            }

            st.write("Step 3/5 — Researching trusted football websites and comparing the real match context.")
            result = generate_explanation(
                match=selected_match,
                bookmaker=selected_bookmaker,
                model_name=model_name,
                system_prompt=system_prompt,
                allowed_domains=allowed_domains,
                max_tool_calls=max_tool_calls,
            )

            st.write("Step 4/5 — Checking the output structure and preparing consistent analysis blocks.")
            st.write("Step 5/5 — Rendering the final explanation on the page.")
            status.update(label="Explanation ready", state="complete", expanded=False)

        st.session_state["last_result"] = result
        st.session_state["last_result_key"] = current_result_key

    except Exception as e:
        st.error(f"Error: {e}")

result_to_show = None
if st.session_state["last_result"] is not None and st.session_state["last_result_key"] == current_result_key:
    result_to_show = st.session_state["last_result"]

if result_to_show:
    st.markdown("## Expert analysis")

    general_text = append_bookmaker_note(
        result_to_show["general_match_description"]["text"],
        make_bookmaker_note("general_match_description", selected_bookmaker, selected_match),
    )

    outcome_text = append_bookmaker_note(
        result_to_show["match_outcome_probability"]["text"],
        make_bookmaker_note("match_outcome_probability", selected_bookmaker, selected_match),
    )

    score_text = append_bookmaker_note(
        result_to_show["correct_score_probability"]["text"],
        make_bookmaker_note("correct_score_probability", selected_bookmaker, selected_match),
    )

    btts_text = append_bookmaker_note(
        result_to_show["both_teams_to_score"]["text"],
        make_bookmaker_note("both_teams_to_score", selected_bookmaker, selected_match),
    )

    goals_text = append_bookmaker_note(
        result_to_show["match_goals_probability"]["text"],
        make_bookmaker_note("match_goals_probability", selected_bookmaker, selected_match),
    )

    render_analysis_block(
        title=result_to_show["general_match_description"]["title"],
        text=general_text,
        risk_note=result_to_show["general_match_description"]["risk_note"],
    )

    render_analysis_block(
        title=result_to_show["match_outcome_probability"]["title"],
        text=outcome_text,
        badge_text=f'Favored outcome: {result_to_show["match_outcome_probability"]["favored_outcome"]}',
    )

    render_analysis_block(
        title=result_to_show["correct_score_probability"]["title"],
        text=score_text,
        badge_text=f'Most likely score: {result_to_show["correct_score_probability"]["most_likely_score"]}',
    )

    render_analysis_block(
        title=result_to_show["both_teams_to_score"]["title"],
        text=btts_text,
        badge_text=f'Most likely BTTS outcome: {result_to_show["both_teams_to_score"]["most_likely_outcome"]}',
    )

    render_analysis_block(
        title=result_to_show["match_goals_probability"]["title"],
        text=goals_text,
    )
